#!/usr/bin/env python3
"""Converte a planilha de precedentes curados em CSV para a esteira."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path


CABECALHO_ESPERADO = {
    "ID da decisão", "Processo", "Tema", "Tribunal", "Disciplina", "Estado jurisprudencial"
}


def carregar_openpyxl():
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Instale 'openpyxl' antes de preparar os itens da esteira.") from exc
    return load_workbook


def ids_normalizados(valores: list[str] | None) -> set[str]:
    return {str(valor).strip() for valor in valores or [] if str(valor).strip()}


def texto(valor) -> str:
    return "" if valor is None else str(valor).strip()


def ler_precedentes(caminho: Path) -> list[dict[str, str]]:
    load_workbook = carregar_openpyxl()
    wb = load_workbook(caminho, read_only=True, data_only=True)
    try:
        if "Precedentes" not in wb.sheetnames:
            raise ValueError("A planilha não contém a aba 'Precedentes'.")
        ws = wb["Precedentes"]
        linhas = ws.iter_rows(values_only=True)
        cabecalho = [texto(valor) for valor in next(linhas, ())]
        faltantes = CABECALHO_ESPERADO.difference(cabecalho)
        if faltantes:
            nomes = ", ".join(sorted(faltantes))
            raise ValueError(f"Cabeçalho incompatível; campos ausentes: {nomes}.")
        indice = {nome: cabecalho.index(nome) for nome in CABECALHO_ESPERADO}
        registros = []
        for numero, linha in enumerate(linhas, start=2):
            valores = list(linha)
            processo = texto(valores[indice["Processo"]] if indice["Processo"] < len(valores) else "")
            if not processo:
                continue
            tribunal = texto(valores[indice["Tribunal"]] if indice["Tribunal"] < len(valores) else "").upper()
            if tribunal not in {"STF", "STJ"}:
                raise ValueError(f"Linha {numero}: tribunal deve ser STF ou STJ.")
            identificador = texto(valores[indice["ID da decisão"]] if indice["ID da decisão"] < len(valores) else "") or processo
            registros.append({
                "id": identificador,
                "tema": texto(valores[indice["Tema"]] if indice["Tema"] < len(valores) else "") or processo,
                "tribunal": tribunal,
                "disciplina": texto(valores[indice["Disciplina"]] if indice["Disciplina"] < len(valores) else ""),
                "estado": texto(valores[indice["Estado jurisprudencial"]] if indice["Estado jurisprudencial"] < len(valores) else "").lower(),
            })
        return registros
    finally:
        wb.close()


def converter(registros: list[dict[str, str]], altas: set[str], erros: set[str], incluir_superados: bool) -> tuple[list[dict[str, str]], int]:
    itens = []
    superados_excluidos = 0
    vistos: set[str] = set()
    for registro in registros:
        identificador = registro["id"]
        if identificador in vistos:
            continue
        vistos.add(identificador)
        if registro["estado"] == "superado" and not incluir_superados:
            superados_excluidos += 1
            continue
        itens.append({
            "id": identificador,
            "tema": registro["tema"],
            "tribunal": registro["tribunal"],
            "disciplina": registro["disciplina"],
            "prioridade": "alta" if identificador in altas else "padrao",
            "origem_erro": "sim" if identificador in erros else "nao",
        })
    return itens, superados_excluidos


def escrever_csv(itens: list[dict[str, str]], saida: Path) -> None:
    saida.parent.mkdir(parents=True, exist_ok=True)
    with saida.open("w", encoding="utf-8", newline="") as arquivo:
        campos = ["id", "tema", "tribunal", "disciplina", "prioridade", "origem_erro"]
        escritor = csv.DictWriter(arquivo, fieldnames=campos)
        escritor.writeheader()
        escritor.writerows(itens)


def main() -> None:
    parser = argparse.ArgumentParser(description="Converte precedentes curados em CSV para a esteira.")
    parser.add_argument("entrada", type=Path, help="Planilha XLSX gerada pela curadoria.")
    parser.add_argument("saida", type=Path, help="CSV para usar em init ou add da esteira.")
    parser.add_argument("--prioridade-alta", action="append", default=[], metavar="ID", help="ID de precedente que deve receber prioridade alta; pode ser repetido.")
    parser.add_argument("--origem-erro", action="append", default=[], metavar="ID", help="ID ligado a erro documentado; pode ser repetido.")
    parser.add_argument("--incluir-superados", action="store_true", help="Inclui registros marcados como superados.")
    args = parser.parse_args()
    if not args.entrada.exists():
        raise SystemExit(f"Arquivo não encontrado: {args.entrada}")
    try:
        registros = ler_precedentes(args.entrada)
        itens, excluidos = converter(registros, ids_normalizados(args.prioridade_alta), ids_normalizados(args.origem_erro), args.incluir_superados)
        escrever_csv(itens, args.saida)
    except (RuntimeError, ValueError) as exc:
        raise SystemExit(f"ERRO: {exc}")
    print(json.dumps({"saida": str(args.saida), "itens": len(itens), "superados_excluidos": excluidos}, ensure_ascii=False))


if __name__ == "__main__":
    main()
