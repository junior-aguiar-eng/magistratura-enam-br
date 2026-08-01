#!/usr/bin/env python3
"""
Motor da esteira de jurisprudência STF/STJ.

Uma esteira contínua de estudo de julgados com revisão espaçada, dois estoques
(Entrada e Revisão) rodando em paralelo, dois regimes (esteira e consolidação) e
diagnóstico honesto de vazão. Determinístico: sem aleatoriedade em nenhuma fila.

Fluxo de uso:
  Execute na raiz do plugin:
  1. Gerar a planilha vazia (ou a partir de uma lista de julgados curados):
       uv run python skills/planejar-jurisprudencia/scripts/atualizar_esteira.py init --prova 2025-11-15 --saida esteira.xlsx
  2. A cada semana: abrir a planilha, marcar o que foi feito na coluna "Feito?"
     (X ou "sim"), salvar, e rodar:
       uv run python skills/planejar-jurisprudencia/scripts/atualizar_esteira.py atualizar --arquivo esteira.xlsx
     O script recalcula datas de revisão, promove itens entre filas e regenera a
     aba Semana.
  3. Adicionar julgados novos (vindos da curadoria) sem recriar a planilha:
       uv run python skills/planejar-jurisprudencia/scripts/atualizar_esteira.py add --arquivo esteira.xlsx --itens novos.csv

O CSV de entrada de julgados novos deve ter as colunas:
  id, tema, tribunal, disciplina, prioridade, origem_erro
onde:
  - prioridade ∈ {alta, padrao}
  - origem_erro ∈ {sim, nao}  (ligado a erro documentado do candidato)

Regras do motor (ver SKILL.md para o racional):
  - Revisão vencida ocupa slot antes de conteúdo novo.
  - Fila de entrada ordenada por prioridade; desempate: origem_erro primeiro.
  - Fila de revisão ordenada por data de vencimento (mais atrasada primeiro).
  - Ciclos por prioridade:
        alta   -> 4 revisões: D+3, D+10, D+30, D+75
        padrao -> 3 revisões: D+5, D+21, D+60
  - ~20% de folga estrutural por semana (a aba Semana nunca preenche 100% da
    capacidade declarada).
  - Regime consolidação: nos últimos DIAS_CONSOLIDACAO dias antes da prova, a
    entrada congela e a Semana prioriza varredura final da prioridade alta.
"""

import argparse
import csv
import datetime as dt
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "openpyxl não encontrado. Na raiz do plugin, execute: uv sync --all-groups\n"
    )
    sys.exit(1)

# ----------------------------------------------------------------------------
# Parâmetros do motor
# ----------------------------------------------------------------------------

CICLOS = {
    "alta": [3, 10, 30, 75],
    "padrao": [5, 21, 60],
}

TRIBUNAIS = {"STF", "STJ"}
ESTADOS_JURISPRUDENCIAIS = {
    "confirmado",
    "confirmado com ressalvas",
    "restringido ou distinguido",
    "superado",
    "controvertido",
    "não verificado",
}
GRAUS_CONFIANCA = {"alto", "médio", "baixo"}
ORIGENS_ERRO = {"sim", "nao"}
COLUNAS_CSV_ITENS = [
    "id", "tema", "tribunal", "disciplina", "estado_jurisprudencial",
    "grau_confianca", "fontes_essenciais", "prioridade",
    "motivo_prioridade", "origem_erro",
]

FOLGA = 0.20            # 20% da capacidade semanal fica livre como amortecedor
DIAS_CONSOLIDACAO = 15  # janela final: entrada congela, varredura de prioridade alta
FUSO_BRASIL = dt.timezone(dt.timedelta(hours=-3))

# Minutos estimados por atividade (usados só para orçamento da aba Semana).
# Primeira passada custa mais que revisão; jurisprudência revisa rápido depois da 1ª leitura.
MIN_PRIMEIRA = 25
MIN_REVISAO = 10

ABAS = ("Entrada", "Revisao", "Remediacao", "Semana", "Config")

# ----------------------------------------------------------------------------
# Estilos
# ----------------------------------------------------------------------------

FONTE = "Arial"
COR_STF = "DDEBF7"   # azul claro
COR_STJ = "E2EFDA"   # verde claro
COR_ALTA = "FCE4D6"  # laranja claro (prioridade alta)
COR_HEADER = "1F4E78"
COR_VENCIDA = "FFC7CE"  # vermelho claro (revisão atrasada)
BORDA = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _header_font():
    return Font(name=FONTE, bold=True, color="FFFFFF", size=11)


def _cell_font(bold=False):
    return Font(name=FONTE, bold=bold, size=10)


def _style_header_row(ws, ncols):
    fill = PatternFill("solid", fgColor=COR_HEADER)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _header_font()
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDA
    ws.freeze_panes = "A2"


# ----------------------------------------------------------------------------
# Helpers de data
# ----------------------------------------------------------------------------

def hoje():
    return dt.datetime.now(FUSO_BRASIL).date()


def parse_data(s):
    if isinstance(s, dt.date):
        return s
    if isinstance(s, dt.datetime):
        return s.date()
    if not s:
        return None
    try:
        return dt.date.fromisoformat(str(s).strip()[:10])
    except ValueError:
        return None  # valores como "—" ou "sprint" não são datas


def fmt(d):
    return d.strftime("%Y-%m-%d") if d else ""


def is_feito(v):
    if v is None:
        return False
    return str(v).strip().lower() in {"x", "sim", "s", "1", "true", "ok", "feito"}


def ler_ciclo(it, total_ciclos):
    """Lê e valida o ciclo de uma revisão antes de avançá-la."""
    valor = it.get("ciclo")
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return 1
    if isinstance(valor, bool) or (isinstance(valor, float) and not valor.is_integer()):
        ciclo = None
    else:
        try:
            ciclo = int(valor)
        except (TypeError, ValueError):
            ciclo = None
    if ciclo is None or not 1 <= ciclo <= total_ciclos:
        identificador = it.get("id") or "sem id"
        raise ValueError(
            f"Ciclo inválido na revisão '{identificador}': esperado um inteiro entre "
            f"1 e {total_ciclos}, encontrado {valor!r}. Corrija a coluna 'ciclo' e execute novamente."
        )
    return ciclo


# ----------------------------------------------------------------------------
# Estrutura de colunas
# ----------------------------------------------------------------------------

COLS_ENTRADA = ["id", "tema", "tribunal", "disciplina", "estado_jurisprudencial",
                "grau_confianca", "fontes_essenciais", "prioridade", "motivo_prioridade", "origem_erro",
                "data_entrada", "Feito?"]
COLS_REVISAO = ["id", "tema", "tribunal", "disciplina", "estado_jurisprudencial",
                "grau_confianca", "fontes_essenciais", "prioridade", "motivo_prioridade", "origem_erro",
                "ciclo", "total_ciclos", "proxima_revisao", "resultado_revisao", "encaminhamento", "Feito?"]
COLS_REMEDIACAO = ["id", "tema", "tribunal", "disciplina", "resultado_revisao", "encaminhamento", "data_registro", "Feito?"]
COLS_SEMANA = ["ordem", "tipo", "id", "tema", "tribunal", "disciplina",
               "estado_jurisprudencial", "grau_confianca", "fontes_essenciais",
               "prioridade", "motivo_prioridade", "vencimento", "min_est", "Feito?"]
COLS_CONFIG = ["chave", "valor"]


# ----------------------------------------------------------------------------
# Leitura / escrita da planilha
# ----------------------------------------------------------------------------

def ler_aba(ws, cols):
    cabecalho = {
        str(ws.cell(row=1, column=coluna).value).strip(): coluna
        for coluna in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=coluna).value is not None
    }
    linhas = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=cabecalho[coluna]).value if coluna in cabecalho else "" for coluna in cols]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        linhas.append(dict(zip(cols, vals)))
    return linhas


def escrever_aba(ws, cols, linhas):
    ws.delete_rows(1, ws.max_row if ws.max_row else 1)
    for i, c in enumerate(cols, start=1):
        ws.cell(row=1, column=i, value=c)
    _style_header_row(ws, len(cols))
    for ri, item in enumerate(linhas, start=2):
        prio = str(item.get("prioridade", "")).lower()
        trib = str(item.get("tribunal", "")).upper()
        for ci, c in enumerate(cols, start=1):
            cell = ws.cell(row=ri, column=ci, value=item.get(c, ""))
            cell.font = _cell_font()
            cell.border = BORDA
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(c in {"tema", "fontes_essenciais"}))
        # realce por tribunal na célula do tribunal
        tcol = cols.index("tribunal") + 1 if "tribunal" in cols else None
        if tcol:
            fill = COR_STF if trib == "STF" else (COR_STJ if trib == "STJ" else None)
            if fill:
                ws.cell(row=ri, column=tcol).fill = PatternFill("solid", fgColor=fill)
        # realce da linha inteira se prioridade alta (coluna prioridade)
        pcol = cols.index("prioridade") + 1 if "prioridade" in cols else None
        if pcol and prio == "alta":
            ws.cell(row=ri, column=pcol).fill = PatternFill("solid", fgColor=COR_ALTA)
        # realce de vencimento atrasado na aba Semana / Revisao
        if "vencimento" in cols or "proxima_revisao" in cols:
            dcol_name = "vencimento" if "vencimento" in cols else "proxima_revisao"
            dval = parse_data(item.get(dcol_name))
            if dval and dval < hoje():
                dc = cols.index(dcol_name) + 1
                ws.cell(row=ri, column=dc).fill = PatternFill("solid", fgColor=COR_VENCIDA)
    _ajustar_larguras(ws, cols)


def _ajustar_larguras(ws, cols):
    larguras = {
        "id": 16, "tema": 46, "tribunal": 10, "disciplina": 20,
        "estado_jurisprudencial": 26, "grau_confianca": 18, "fontes_essenciais": 48,
        "prioridade": 12, "motivo_prioridade": 28, "origem_erro": 12, "data_entrada": 14,
        "ciclo": 8, "total_ciclos": 12, "proxima_revisao": 16,
        "proxima_data": 16, "vencimento": 14, "min_est": 9,
        "ordem": 8, "tipo": 12, "Feito?": 9, "chave": 24, "valor": 30,
    }
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(c, 16)


def ler_config(ws):
    cfg = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k:
            cfg[str(k).strip()] = v
    return cfg


def escrever_config(ws, cfg):
    ws.delete_rows(1, ws.max_row if ws.max_row else 1)
    ws.cell(row=1, column=1, value="chave")
    ws.cell(row=1, column=2, value="valor")
    _style_header_row(ws, 2)
    for ri, (k, v) in enumerate(cfg.items(), start=2):
        ws.cell(row=ri, column=1, value=k).font = _cell_font(bold=True)
        ws.cell(row=ri, column=2, value=v).font = _cell_font()
    _ajustar_larguras(ws, COLS_CONFIG)


# ----------------------------------------------------------------------------
# Capacidade semanal
# ----------------------------------------------------------------------------

DIAS_SEMANA = ["seg", "ter", "qua", "qui", "sex", "sab", "dom"]
CAP_PADRAO = {"seg": 60, "ter": 60, "qua": 60, "qui": 60, "sex": 60,
              "sab": 180, "dom": 180}  # minutos/dia


def capacidade_semanal_min(cfg):
    total = 0
    for d in DIAS_SEMANA:
        try:
            total += int(cfg.get(f"cap_{d}", CAP_PADRAO[d]))
        except (TypeError, ValueError):
            total += CAP_PADRAO[d]
    return total


# ----------------------------------------------------------------------------
# init
# ----------------------------------------------------------------------------

def cmd_init(args):
    if args.prova and not parse_data(args.prova):
        raise ValueError("A data da prova deve usar AAAA-MM-DD.")
    wb = Workbook()
    wb.remove(wb.active)
    for nome in ABAS:
        wb.create_sheet(nome)

    escrever_aba(wb["Entrada"], COLS_ENTRADA, [])
    escrever_aba(wb["Revisao"], COLS_REVISAO, [])
    escrever_aba(wb["Remediacao"], COLS_REMEDIACAO, [])
    escrever_aba(wb["Semana"], COLS_SEMANA, [])

    cfg = {
        "prova": args.prova or "",
        "regime": "esteira",
        "criado_em": fmt(hoje()),
        "atualizado_em": fmt(hoje()),
    }
    for d in DIAS_SEMANA:
        cfg[f"cap_{d}"] = CAP_PADRAO[d]
    escrever_config(wb["Config"], cfg)

    novos = []

    # Se veio lista de julgados, carrega direto na Entrada
    if args.itens:
        novos = _ler_csv_itens(args.itens)
        for it in novos:
            it["data_entrada"] = fmt(hoje())
            it["Feito?"] = ""
        escrever_aba(wb["Entrada"], COLS_ENTRADA, novos)

    Path(args.saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.saida)
    print(json.dumps({
        "status": "criado",
        "arquivo": args.saida,
        "prova": args.prova or "não definida",
        "itens_iniciais": len(novos),
        "dica": "Rode 'atualizar' após marcar a coluna Feito? para gerar a aba Semana.",
    }, ensure_ascii=False, indent=2))


def _ler_csv_itens(caminho):
    itens = []
    identificadores = set()
    with open(caminho, newline="", encoding="utf-8") as f:
        leitor = csv.DictReader(f)
        cabecalho = set(leitor.fieldnames or [])
        faltantes = [campo for campo in COLUNAS_CSV_ITENS if campo not in cabecalho]
        if faltantes:
            raise ValueError(f"CSV de itens sem colunas obrigatórias: {', '.join(faltantes)}.")
        for numero, row in enumerate(leitor, start=2):
            item = {campo: (row.get(campo) or "").strip() for campo in COLUNAS_CSV_ITENS}
            item["tribunal"] = item["tribunal"].upper()
            item["estado_jurisprudencial"] = item["estado_jurisprudencial"].lower()
            item["grau_confianca"] = item["grau_confianca"].lower()
            if item["grau_confianca"] == "medio":
                item["grau_confianca"] = "médio"
            item["prioridade"] = item["prioridade"].lower()
            item["origem_erro"] = item["origem_erro"].lower()

            obrigatorios = [campo for campo in COLUNAS_CSV_ITENS if not item[campo]]
            if obrigatorios:
                raise ValueError(f"Linha {numero}: campos obrigatórios vazios: {', '.join(obrigatorios)}.")
            if item["id"] in identificadores:
                raise ValueError(f"Linha {numero}: id duplicado no CSV: {item['id']}.")
            if item["tribunal"] not in TRIBUNAIS:
                raise ValueError(f"Linha {numero}: tribunal deve ser STF ou STJ.")
            if item["estado_jurisprudencial"] not in ESTADOS_JURISPRUDENCIAIS:
                raise ValueError(f"Linha {numero}: estado jurisprudencial inválido: {item['estado_jurisprudencial']}.")
            if item["grau_confianca"] not in GRAUS_CONFIANCA:
                raise ValueError(f"Linha {numero}: grau de confiança inválido: {item['grau_confianca']}.")
            if item["prioridade"] not in CICLOS:
                raise ValueError(f"Linha {numero}: prioridade deve ser alta ou padrao.")
            if item["origem_erro"] not in ORIGENS_ERRO:
                raise ValueError(f"Linha {numero}: origem_erro deve ser sim ou nao.")
            if item["prioridade"] == "alta" and item["motivo_prioridade"] == "prioridade padrão":
                raise ValueError(f"Linha {numero}: prioridade alta exige motivo específico.")

            identificadores.add(item["id"])
            itens.append(item)
    return itens


# ----------------------------------------------------------------------------
# add
# ----------------------------------------------------------------------------

def cmd_add(args):
    wb = load_workbook(args.arquivo)
    cfg = ler_config(wb["Config"])
    if _em_consolidacao(cfg):
        print(json.dumps({
            "status": "bloqueado",
            "motivo": "Regime de consolidação ativo — entrada de conteúdo novo está congelada.",
            "acao": "Foque na aba Semana (varredura final). Novos julgados só após a prova.",
        }, ensure_ascii=False, indent=2))
        return

    entrada = ler_aba(wb["Entrada"], COLS_ENTRADA)
    existentes = {str(i["id"]).strip() for i in entrada}
    # também não readicionar o que já está em revisão
    for i in ler_aba(wb["Revisao"], COLS_REVISAO):
        existentes.add(str(i["id"]).strip())

    novos = _ler_csv_itens(args.itens)
    add = 0
    for it in novos:
        if it["id"] in existentes:
            continue
        it["data_entrada"] = fmt(hoje())
        it["Feito?"] = ""
        entrada.append(it)
        existentes.add(it["id"])
        add += 1

    escrever_aba(wb["Entrada"], COLS_ENTRADA, entrada)
    cfg["atualizado_em"] = fmt(hoje())
    escrever_config(wb["Config"], cfg)
    wb.save(args.arquivo)
    print(json.dumps({
        "status": "adicionado",
        "novos": add,
        "ignorados_duplicados": len(novos) - add,
        "total_entrada": len(entrada),
    }, ensure_ascii=False, indent=2))


# ----------------------------------------------------------------------------
# atualizar  — o coração do motor
# ----------------------------------------------------------------------------

def _prio_rank(item):
    # ordem de prioridade para a fila de entrada: alta antes de padrão;
    # dentro do mesmo nível, origem_erro (gap documentado) primeiro.
    prio = str(item.get("prioridade", "padrao")).lower()
    erro = str(item.get("origem_erro", "nao")).lower() in {"sim", "s", "1", "true"}
    base = 0 if prio == "alta" else 1
    return (base, 0 if erro else 1)


def _em_consolidacao(cfg):
    prova = parse_data(cfg.get("prova"))
    if not prova:
        return False
    return (prova - hoje()).days <= DIAS_CONSOLIDACAO and (prova - hoje()).days >= -1


def cmd_atualizar(args):
    wb = load_workbook(args.arquivo)
    if "Remediacao" not in wb.sheetnames:
        wb.create_sheet("Remediacao")
        escrever_aba(wb["Remediacao"], COLS_REMEDIACAO, [])
    cfg = ler_config(wb["Config"])
    entrada = ler_aba(wb["Entrada"], COLS_ENTRADA)
    revisao = ler_aba(wb["Revisao"], COLS_REVISAO)
    remediacao = ler_aba(wb["Remediacao"], COLS_REMEDIACAO)

    hoje_d = hoje()
    consolidacao = _em_consolidacao(cfg)
    cfg["regime"] = "consolidacao" if consolidacao else "esteira"

    # ---- 1. Processar itens de ENTRADA marcados como feitos -> promover a REVISÃO
    promovidos = 0
    entrada_restante = []
    for it in entrada:
        if is_feito(it.get("Feito?")):
            prio = str(it.get("prioridade", "padrao")).lower()
            ciclos = CICLOS.get(prio, CICLOS["padrao"])
            revisao.append({
                "id": it["id"],
                "tema": it["tema"],
                "tribunal": it["tribunal"],
                "disciplina": it["disciplina"],
                "estado_jurisprudencial": it.get("estado_jurisprudencial", ""),
                "grau_confianca": it.get("grau_confianca", ""),
                "fontes_essenciais": it.get("fontes_essenciais", ""),
                "prioridade": prio,
                "motivo_prioridade": it.get("motivo_prioridade", "prioridade padrão"),
                "origem_erro": it.get("origem_erro", "nao"),
                "ciclo": 1,
                "total_ciclos": len(ciclos),
                "proxima_revisao": fmt(hoje_d + dt.timedelta(days=ciclos[0])),
                "resultado_revisao": "",
                "encaminhamento": "",
                "Feito?": "",
            })
            promovidos += 1
        else:
            entrada_restante.append(it)

    # ---- 2. Processar REVISÕES marcadas como feitas -> avançar ciclo ou concluir
    concluidos = 0
    avancados = 0
    revisao_nova = []
    for it in revisao:
        if is_feito(it.get("Feito?")):
            resultado = str(it.get("resultado_revisao", "")).strip().lower()
            encaminhamento = str(it.get("encaminhamento", "")).strip().lower()
            if resultado in {"erro", "revisar"} and encaminhamento in {"questao_objetiva", "discursiva_curta", "prova_oral"}:
                remediacao.append({
                    "id": it["id"], "tema": it["tema"], "tribunal": it["tribunal"],
                    "disciplina": it["disciplina"], "resultado_revisao": resultado,
                    "encaminhamento": encaminhamento, "data_registro": fmt(hoje_d), "Feito?": "",
                })
            prio = str(it.get("prioridade", "padrao")).lower()
            ciclos = CICLOS.get(prio, CICLOS["padrao"])
            ciclo_atual = ler_ciclo(it, len(ciclos))
            if ciclo_atual >= len(ciclos):
                concluidos += 1  # consolidado; sai da esteira
                continue
            prox = ciclos[ciclo_atual]  # próximo offset a partir de hoje
            it["ciclo"] = ciclo_atual + 1
            it["proxima_revisao"] = fmt(hoje_d + dt.timedelta(days=prox))
            it["Feito?"] = ""
            it["resultado_revisao"] = ""
            it["encaminhamento"] = ""
            revisao_nova.append(it)
            avancados += 1
        else:
            it["Feito?"] = ""
            revisao_nova.append(it)

    revisao = revisao_nova

    escrever_aba(wb["Entrada"], COLS_ENTRADA, entrada_restante)
    escrever_aba(wb["Revisao"], COLS_REVISAO, revisao)
    escrever_aba(wb["Remediacao"], COLS_REMEDIACAO, remediacao)

    # ---- 3. Gerar aba SEMANA
    semana, diag = _montar_semana(entrada_restante, revisao, cfg, consolidacao)
    escrever_aba(wb["Semana"], COLS_SEMANA, semana)

    cfg["atualizado_em"] = fmt(hoje_d)
    escrever_config(wb["Config"], cfg)
    wb.save(args.arquivo)

    saida = {
        "status": "atualizado",
        "regime": cfg["regime"],
        "promovidos_para_revisao": promovidos,
        "revisoes_avancadas": avancados,
        "consolidados": concluidos,
        "fila_entrada": len(entrada_restante),
        "fila_revisao": len(revisao),
        "itens_na_semana": len(semana),
        "diagnostico": diag,
    }
    print(json.dumps(saida, ensure_ascii=False, indent=2))


def _montar_semana(entrada, revisao, cfg, consolidacao):
    """
    Monta a lista da semana respeitando a capacidade e as regras de prioridade.
    Retorna (linhas_semana, diagnostico).
    """
    hoje_d = hoje()
    cap_total = capacidade_semanal_min(cfg)
    orcamento = int(cap_total * (1 - FOLGA))  # aplica folga estrutural

    linhas = []
    usado = 0
    ordem = 1

    # -- Fila de revisão: ordenada por vencimento (mais atrasada primeiro).
    revs = []
    for it in revisao:
        venc = parse_data(it.get("proxima_revisao"))
        # janela da semana: vence de hoje até +7, OU já venceu (atrasada)
        if venc and venc <= hoje_d + dt.timedelta(days=7):
            revs.append((venc, it))
    revs.sort(key=lambda x: x[0])  # data crescente = mais atrasada primeiro

    for venc, it in revs:
        if usado + MIN_REVISAO > orcamento:
            break
        linhas.append({
            "ordem": ordem, "tipo": "revisão", "id": it["id"], "tema": it["tema"],
            "tribunal": it["tribunal"], "disciplina": it["disciplina"],
            "estado_jurisprudencial": it.get("estado_jurisprudencial", ""),
            "grau_confianca": it.get("grau_confianca", ""),
            "fontes_essenciais": it.get("fontes_essenciais", ""),
            "prioridade": it["prioridade"], "vencimento": fmt(venc),
            "motivo_prioridade": it.get("motivo_prioridade", "prioridade padrão"),
            "min_est": MIN_REVISAO, "Feito?": "",
        })
        usado += MIN_REVISAO
        ordem += 1

    # -- Fila de entrada.
    if consolidacao:
        # Em consolidação: não entra conteúdo novo. Em vez disso, varredura final
        # da prioridade alta que ainda está em revisão mas não coube acima.
        alta_restante = [it for it in revisao
                         if str(it.get("prioridade")).lower() == "alta"
                         and it["id"] not in {l["id"] for l in linhas}]
        alta_restante.sort(key=_prio_rank)
        for it in alta_restante:
            if usado + MIN_REVISAO > orcamento:
                break
            linhas.append({
                "ordem": ordem, "tipo": "varredura-final", "id": it["id"],
                "tema": it["tema"], "tribunal": it["tribunal"],
                "disciplina": it["disciplina"],
                "estado_jurisprudencial": it.get("estado_jurisprudencial", ""),
                "grau_confianca": it.get("grau_confianca", ""),
                "fontes_essenciais": it.get("fontes_essenciais", ""),
                "prioridade": it["prioridade"],
                "motivo_prioridade": it.get("motivo_prioridade", "prioridade padrão"),
                "vencimento": "sprint", "min_est": MIN_REVISAO, "Feito?": "",
            })
            usado += MIN_REVISAO
            ordem += 1
    else:
        entrada_ord = sorted(entrada, key=_prio_rank)
        for it in entrada_ord:
            if usado + MIN_PRIMEIRA > orcamento:
                break
            linhas.append({
                "ordem": ordem, "tipo": "novo", "id": it["id"], "tema": it["tema"],
                "tribunal": it["tribunal"], "disciplina": it["disciplina"],
                "estado_jurisprudencial": it.get("estado_jurisprudencial", ""),
                "grau_confianca": it.get("grau_confianca", ""),
                "fontes_essenciais": it.get("fontes_essenciais", ""),
                "prioridade": it["prioridade"], "vencimento": "—",
                "motivo_prioridade": it.get("motivo_prioridade", "prioridade padrão"),
                "min_est": MIN_PRIMEIRA, "Feito?": "",
            })
            usado += MIN_PRIMEIRA
            ordem += 1

    diag = _diagnostico(entrada, revisao, cap_total, orcamento, usado, revs, cfg, consolidacao)
    return linhas, diag


def _diagnostico(entrada, revisao, cap_total, orcamento, usado, revs, cfg, consolidacao):
    """Diagnóstico honesto de vazão: a esteira comporta o volume atual?"""
    hoje_d = hoje()
    atrasadas = sum(1 for venc, _ in revs if venc < hoje_d)
    revisao_semana_min = sum(1 for venc, _ in revs) * MIN_REVISAO

    alertas = []

    # Revisão sozinha já estoura o orçamento? sinal vermelho.
    if revisao_semana_min > orcamento:
        alertas.append(
            "A carga de REVISÃO desta semana já excede sua capacidade útil. "
            "Nenhum conteúdo novo deveria entrar até você drenar a revisão. "
            "Considere reduzir a seleção da curadoria ou aumentar a capacidade semanal."
        )
    if atrasadas > 0:
        alertas.append(
            f"{atrasadas} revisão(ões) já venceram — elas têm prioridade absoluta "
            "sobre conteúdo novo nesta semana."
        )

    # Vazão de entrada: quantas semanas para drenar a fila de entrada atual?
    folga_para_novos = max(orcamento - revisao_semana_min, 0)
    novos_por_semana = folga_para_novos // MIN_PRIMEIRA if MIN_PRIMEIRA else 0
    semanas_para_drenar = (
        round(len(entrada) / novos_por_semana, 1) if novos_por_semana else None
    )

    prova = parse_data(cfg.get("prova"))
    semanas_ate_prova = None
    if prova:
        semanas_ate_prova = max((prova - hoje_d).days // 7, 0)

    if semanas_para_drenar is None and len(entrada) > 0 and not consolidacao:
        alertas.append(
            "Vazão de entrada ~zero: a revisão consome toda a capacidade útil. "
            "A fila de entrada não anda nesta configuração."
        )
    elif (semanas_para_drenar and semanas_ate_prova
          and semanas_para_drenar > semanas_ate_prova):
        alertas.append(
            f"No ritmo atual, drenar a fila de entrada levaria ~{semanas_para_drenar} "
            f"semanas, mas restam ~{semanas_ate_prova} até a prova. "
            "Priorize: nem todo julgado de 2025 caberá — foque no gap documentado "
            "(origem_erro=sim) e na prioridade alta."
        )

    return {
        "regime": "consolidacao" if consolidacao else "esteira",
        "capacidade_semanal_min": cap_total,
        "orcamento_util_min": orcamento,
        "usado_na_semana_min": usado,
        "revisoes_na_janela": len(revs),
        "revisoes_atrasadas": atrasadas,
        "fila_entrada": len(entrada),
        "novos_por_semana_estimado": int(novos_por_semana),
        "semanas_para_drenar_entrada": semanas_para_drenar,
        "semanas_ate_prova": semanas_ate_prova,
        "alertas": alertas or ["Vazão saudável: a esteira comporta o volume atual."],
    }


# ----------------------------------------------------------------------------
# status  — leitura rápida sem alterar nada
# ----------------------------------------------------------------------------

def cmd_status(args):
    wb = load_workbook(args.arquivo)
    cfg = ler_config(wb["Config"])
    entrada = ler_aba(wb["Entrada"], COLS_ENTRADA)
    revisao = ler_aba(wb["Revisao"], COLS_REVISAO)
    hoje_d = hoje()
    atrasadas = 0
    for it in revisao:
        v = parse_data(it.get("proxima_revisao"))
        if v and v < hoje_d:
            atrasadas += 1
    print(json.dumps({
        "prova": cfg.get("prova") or "não definida",
        "regime": "consolidacao" if _em_consolidacao(cfg) else "esteira",
        "fila_entrada": len(entrada),
        "fila_revisao": len(revisao),
        "revisoes_atrasadas": atrasadas,
        "atualizado_em": cfg.get("atualizado_em"),
    }, ensure_ascii=False, indent=2))


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description="Motor da esteira de jurisprudência STF/STJ.")
    sub = p.add_subparsers(dest="cmd", required=True)

    pi = sub.add_parser("init", help="Cria a planilha da esteira.")
    pi.add_argument("--prova", help="Data da prova, formato AAAA-MM-DD.")
    pi.add_argument("--itens", help="CSV opcional de julgados iniciais.")
    pi.add_argument("--saida", required=True, help="Caminho do .xlsx de saída.")
    pi.set_defaults(func=cmd_init)

    pa = sub.add_parser("add", help="Adiciona julgados novos à fila de entrada.")
    pa.add_argument("--arquivo", required=True)
    pa.add_argument("--itens", required=True, help="CSV de julgados novos.")
    pa.set_defaults(func=cmd_add)

    pu = sub.add_parser("atualizar", help="Processa 'Feito?', recalcula revisões e regenera a Semana.")
    pu.add_argument("--arquivo", required=True)
    pu.set_defaults(func=cmd_atualizar)

    ps = sub.add_parser("status", help="Leitura rápida do estado, sem alterar a planilha.")
    ps.add_argument("--arquivo", required=True)
    ps.set_defaults(func=cmd_status)

    args = p.parse_args()
    try:
        args.func(args)
    except ValueError as exc:
        raise SystemExit(f"ERRO: {exc}") from exc


if __name__ == "__main__":
    main()
