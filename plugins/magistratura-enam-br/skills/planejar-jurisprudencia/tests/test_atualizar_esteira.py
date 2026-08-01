import csv
from datetime import date
from pathlib import Path
from types import SimpleNamespace

import atualizar_esteira as motor
import pytest
from openpyxl import load_workbook

HOJE = date(2026, 1, 1)


def test_ajuda_do_motor_indica_ambiente_uv():
    codigo = Path(motor.__file__).read_text(encoding="utf-8")
    assert "uv run python skills/planejar-jurisprudencia/scripts/atualizar_esteira.py" in codigo
    assert "uv sync --all-groups" in codigo
    assert "pip install openpyxl" not in codigo


def escrever_itens(caminho: Path, itens: list[dict]) -> None:
    campos = [
        "id", "tema", "tribunal", "disciplina", "estado_jurisprudencial",
        "grau_confianca", "fontes_essenciais", "prioridade",
        "motivo_prioridade", "origem_erro",
    ]
    with caminho.open("w", encoding="utf-8", newline="") as arquivo:
        escritor = csv.DictWriter(arquivo, fieldnames=campos)
        escritor.writeheader()
        escritor.writerows(itens)


def item(identificador="STF-1", prioridade="alta") -> dict:
    return {
        "id": identificador,
        "tema": "Tema de controle",
        "tribunal": "STF",
        "disciplina": "Constitucional",
        "estado_jurisprudencial": "confirmado",
        "grau_confianca": "alto",
        "fontes_essenciais": "acórdão oficial",
        "prioridade": prioridade,
        "motivo_prioridade": "erro documentado" if prioridade == "alta" else "prioridade padrão",
        "origem_erro": "sim" if prioridade == "alta" else "nao",
    }


def carregar_linhas(caminho: Path, aba: str, colunas: list[str]) -> list[dict]:
    wb = load_workbook(caminho)
    try:
        return motor.ler_aba(wb[aba], colunas)
    finally:
        wb.close()


def marcar(caminho: Path, aba: str, valores: dict[str, str]) -> None:
    wb = load_workbook(caminho)
    try:
        ws = wb[aba]
        cabecalho = {celula.value: celula.column for celula in ws[1]}
        for campo, valor in valores.items():
            ws.cell(row=2, column=cabecalho[campo], value=valor)
        wb.save(caminho)
    finally:
        wb.close()


def test_ciclo_alta_promove_revisa_e_encaminha_remediacao(tmp_path, monkeypatch):
    monkeypatch.setattr(motor, "hoje", lambda: HOJE)
    csv_itens = tmp_path / "itens.csv"
    esteira = tmp_path / "esteira.xlsx"
    escrever_itens(csv_itens, [item()])
    motor.cmd_init(SimpleNamespace(prova="2026-11-01", itens=str(csv_itens), saida=str(esteira)))

    marcar(esteira, "Entrada", {"Feito?": "X"})
    motor.cmd_atualizar(SimpleNamespace(arquivo=str(esteira)))
    revisoes = carregar_linhas(esteira, "Revisao", motor.COLS_REVISAO)
    assert revisoes[0]["ciclo"] == 1
    assert revisoes[0]["proxima_revisao"] == "2026-01-04"

    marcar(esteira, "Revisao", {
        "Feito?": "X",
        "resultado_revisao": "erro",
        "encaminhamento": "questao_objetiva",
    })
    motor.cmd_atualizar(SimpleNamespace(arquivo=str(esteira)))
    revisoes = carregar_linhas(esteira, "Revisao", motor.COLS_REVISAO)
    remediacoes = carregar_linhas(esteira, "Remediacao", motor.COLS_REMEDIACAO)
    assert revisoes[0]["ciclo"] == 2
    assert revisoes[0]["proxima_revisao"] == "2026-01-11"
    assert [{chave: linha[chave] for chave in motor.COLS_REMEDIACAO if chave != "Feito?"} for linha in remediacoes] == [{
        "id": "STF-1", "tema": "Tema de controle", "tribunal": "STF",
        "disciplina": "Constitucional", "resultado_revisao": "erro",
        "encaminhamento": "questao_objetiva", "data_registro": "2026-01-01",
    }]


def test_init_le_csv_de_itens_uma_unica_vez(tmp_path, monkeypatch):
    csv_itens = tmp_path / "itens.csv"
    esteira = tmp_path / "esteira.xlsx"
    escrever_itens(csv_itens, [item()])
    leitor_original = motor._ler_csv_itens
    chamadas = []

    def leitor_monitorado(caminho):
        chamadas.append(caminho)
        return leitor_original(caminho)

    monkeypatch.setattr(motor, "_ler_csv_itens", leitor_monitorado)
    motor.cmd_init(SimpleNamespace(prova="2026-11-01", itens=str(csv_itens), saida=str(esteira)))

    assert chamadas == [str(csv_itens)]


@pytest.mark.parametrize("valor", ["texto", 1.5, 0, 5, True])
def test_ler_ciclo_rejeita_valor_invalido(valor):
    with pytest.raises(ValueError, match="Ciclo inválido na revisão 'STF-1'"):
        motor.ler_ciclo({"id": "STF-1", "ciclo": valor}, 4)


def test_add_rejeita_duplicado_repetido_no_mesmo_csv(tmp_path, monkeypatch):
    monkeypatch.setattr(motor, "hoje", lambda: HOJE)
    esteira = tmp_path / "esteira.xlsx"
    csv_itens = tmp_path / "duplicados.csv"
    motor.cmd_init(SimpleNamespace(prova="2026-11-01", itens=None, saida=str(esteira)))
    escrever_itens(csv_itens, [item("STJ-9", "padrao"), item("STJ-9", "padrao")])

    with pytest.raises(ValueError, match="id duplicado no CSV: STJ-9"):
        motor.cmd_add(SimpleNamespace(arquivo=str(esteira), itens=str(csv_itens)))

    entrada = carregar_linhas(esteira, "Entrada", motor.COLS_ENTRADA)
    assert entrada == []


def test_consolidacao_bloqueia_entrada_nova(tmp_path, monkeypatch):
    monkeypatch.setattr(motor, "hoje", lambda: HOJE)
    esteira = tmp_path / "esteira.xlsx"
    csv_itens = tmp_path / "novo.csv"
    motor.cmd_init(SimpleNamespace(prova="2026-01-15", itens=None, saida=str(esteira)))
    escrever_itens(csv_itens, [item("STJ-10", "padrao")])

    motor.cmd_add(SimpleNamespace(arquivo=str(esteira), itens=str(csv_itens)))

    assert carregar_linhas(esteira, "Entrada", motor.COLS_ENTRADA) == []


@pytest.mark.parametrize(
    ("alteracoes", "mensagem"),
    [
        ({"id": ""}, "campos obrigatórios vazios: id"),
        ({"tribunal": "TRF"}, "tribunal deve ser STF ou STJ"),
        ({"estado_jurisprudencial": "incerto"}, "estado jurisprudencial inválido"),
        ({"prioridade": "urgente"}, "prioridade deve ser alta ou padrao"),
        ({"origem_erro": "talvez"}, "origem_erro deve ser sim ou nao"),
    ],
)
def test_csv_rejeita_valores_invalidos_sem_normalizacao_silenciosa(tmp_path, alteracoes, mensagem):
    csv_itens = tmp_path / "invalido.csv"
    dados = item("STJ-11", "padrao")
    dados.update(alteracoes)
    escrever_itens(csv_itens, [dados])

    with pytest.raises(ValueError, match=mensagem):
        motor._ler_csv_itens(csv_itens)


def test_csv_rejeita_id_duplicado_na_criacao_inicial(tmp_path):
    csv_itens = tmp_path / "duplicado.csv"
    escrever_itens(csv_itens, [item("STJ-12", "padrao"), item("STJ-12", "padrao")])

    with pytest.raises(ValueError, match="id duplicado no CSV: STJ-12"):
        motor._ler_csv_itens(csv_itens)


def test_csv_rejeita_coluna_obrigatoria_ausente(tmp_path):
    csv_itens = tmp_path / "sem_tribunal.csv"
    csv_itens.write_text("id,tema\nSTJ-13,Tema de teste\n", encoding="utf-8")

    with pytest.raises(ValueError, match="CSV de itens sem colunas obrigatórias: tribunal"):
        motor._ler_csv_itens(csv_itens)
