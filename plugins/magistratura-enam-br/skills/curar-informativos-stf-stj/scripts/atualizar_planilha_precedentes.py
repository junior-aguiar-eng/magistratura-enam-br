#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path

from common import (
    CONFIANCAS,
    ESTADOS,
    identificador_decisao,
    normalizar_processo,
    sanitizar_excel,
    texto,
    validar_data_iso,
)

CABECALHO = [
    "ID da decisão", "Processo", "Tema", "Tribunal", "Órgão julgador",
    "Disciplina", "Estado jurisprudencial", "Grau de confiança",
    "Data do julgamento", "Data do informativo", "Tese resumida",
    "Fontes essenciais", "Data de inclusão",
]


def validar_registro(item, indice):
    contexto = f"Registro {indice}"
    if not isinstance(item, dict):
        raise ValueError(f"{contexto}: cada item deve ser um objeto JSON.")
    if not normalizar_processo(item.get("processo")):
        raise ValueError(f"{contexto}: campo 'processo' ausente ou inválido.")
    tribunal = texto(item.get("tribunal")).upper()
    if tribunal not in {"STF", "STJ"}:
        raise ValueError(f"{contexto}: 'tribunal' deve ser STF ou STJ.")
    estado = texto(item.get("estado_jurisprudencial")).lower()
    if estado not in ESTADOS:
        raise ValueError(f"{contexto}: estado jurisprudencial inválido: {estado or '[vazio]' }.")
    confianca = texto(item.get("grau_confianca")).lower()
    if confianca not in CONFIANCAS:
        raise ValueError(f"{contexto}: grau de confiança inválido: {confianca or '[vazio]'}.")
    if estado != "confirmado" and not texto(item.get("nota_estado")):
        raise ValueError(f"{contexto}: 'nota_estado' é obrigatória quando o estado não é confirmado.")
    validar_data_iso(item.get("data_julgamento"), "data_julgamento", contexto)
    validar_data_iso(item.get("data_informativo"), "data_informativo", contexto)
    return item


def carregar_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("Instale 'openpyxl' antes de atualizar a planilha.") from exc
    return Workbook, load_workbook, Alignment, Font, get_column_letter, Table, TableStyleInfo


def abrir_ou_criar(caminho, Workbook, load_workbook):
    if caminho.exists():
        wb = load_workbook(caminho)
        ws = wb["Precedentes"] if "Precedentes" in wb.sheetnames else wb.active
        atual = [c.value for c in ws[1]] if ws.max_row else []
        if atual != CABECALHO:
            raise RuntimeError("Cabeçalho incompatível com a versão atual da skill. Use uma cópia migrada ou uma planilha nova.")
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Precedentes"
    ws.append(CABECALHO)
    return wb, ws


def formatar_planilha(ws, Alignment, Font, get_column_letter, Table, TableStyleInfo):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    larguras = [34, 24, 24, 10, 24, 24, 28, 18, 18, 18, 68, 60, 18]
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ref = f"A1:M{max(ws.max_row, 1)}"
    if ws.tables:
        for table in ws.tables.values():
            table.ref = ref
    elif ws.max_row >= 2:
        table = Table(displayName="TabelaPrecedentes", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)


def main():
    parser = argparse.ArgumentParser(description="Acrescenta precedentes validados a uma planilha XLSX.")
    parser.add_argument("novos_julgados", type=Path)
    parser.add_argument("planilha", type=Path)
    parser.add_argument("--saida", type=Path)
    parser.add_argument("--data-inclusao", default=date.today().isoformat())
    args = parser.parse_args()

    if not args.novos_julgados.exists():
        raise SystemExit(f"Arquivo não encontrado: {args.novos_julgados}")
    validar_data_iso(args.data_inclusao, "data-inclusao")
    try:
        dados = json.loads(args.novos_julgados.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"JSON inválido: {exc}")
    if not isinstance(dados, list):
        raise SystemExit("O JSON deve ser uma lista de objetos.")

    try:
        dados = [validar_registro(item, i) for i, item in enumerate(dados, 1)]
        deps = carregar_openpyxl()
        wb, ws = abrir_ou_criar(args.planilha, deps[0], deps[1])
    except (ValueError, RuntimeError) as exc:
        raise SystemExit(f"ERRO: {exc}")

    existentes = {texto(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]}
    adicionados = 0
    duplicados = []
    for item in dados:
        chave = identificador_decisao(item)
        if chave in existentes:
            duplicados.append(texto(item.get("processo")))
            continue
        fontes = item.get("fontes_essenciais", [])
        if isinstance(fontes, list):
            fontes = "\n".join(texto(x) for x in fontes if texto(x))
        linha = [
            chave, item.get("processo"), item.get("tema"), texto(item.get("tribunal")).upper(),
            item.get("orgao_julgador"), item.get("disciplina"), texto(item.get("estado_jurisprudencial")).lower(),
            texto(item.get("grau_confianca")).lower(), item.get("data_julgamento"), item.get("data_informativo"),
            item.get("tese_resumida"), fontes, args.data_inclusao,
        ]
        ws.append([sanitizar_excel(v) for v in linha])
        existentes.add(chave)
        adicionados += 1

    formatar_planilha(ws, *deps[2:])
    saida = args.saida or args.planilha
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)
    print(json.dumps({"saida": str(saida), "adicionados": adicionados, "duplicados": duplicados}, ensure_ascii=False))


if __name__ == "__main__":
    main()
