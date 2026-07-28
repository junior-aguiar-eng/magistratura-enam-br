import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from atualizar_planilha_precedentes import validar_dados, validar_registro


def registro(**kwargs):
    base = {
        "processo": "RE 123/DF",
        "tribunal": "STF",
        "estado_jurisprudencial": "confirmado",
        "grau_confianca": "alto",
        "data_informativo": "2026-07-11",
        "tese_resumida": "Tese de teste",
    }
    base.update(kwargs)
    return base


def test_registro_invalido_rejeitado():
    with pytest.raises(ValueError):
        validar_registro(registro(tribunal="TRF"), 1)


def test_estado_nao_confirmado_exige_nota():
    with pytest.raises(ValueError):
        validar_registro(registro(estado_jurisprudencial="controvertido", grau_confianca="baixo"), 1)


def test_entrada_rejeita_campo_fora_do_contrato():
    with pytest.raises(ValueError, match="Contrato precedentes.schema.json"):
        validar_dados([registro(campo_editorial_solto="não permitido")])


def test_planilha_sanitiza_e_deduplica(tmp_path):
    root = Path(__file__).resolve().parents[1]
    entrada = tmp_path / "dados.json"
    planilha = tmp_path / "precedentes.xlsx"
    dados = [registro(tema='=HYPERLINK("x","y")'), registro()]
    entrada.write_text(json.dumps(dados, ensure_ascii=False), encoding="utf-8")
    subprocess.run([sys.executable, str(root / "scripts" / "atualizar_planilha_precedentes.py"), str(entrada), str(planilha)], check=True)
    wb = openpyxl.load_workbook(planilha, data_only=False)
    ws = wb["Precedentes"]
    assert ws.max_row == 2
    assert str(ws["C2"].value).startswith("'")
    assert ws.freeze_panes == "A2"
